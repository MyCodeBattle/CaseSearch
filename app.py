"""
行政处罚决定书类案检索系统
Flask 主应用
"""
import json
import queue
import threading
from io import BytesIO
from datetime import datetime
from dotenv import load_dotenv
from flask import Flask, render_template, request, jsonify, Response, send_file
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

# 加载环境变量 (.env)
load_dotenv()

from modules.case_loader import load_config, get_available_types, get_cases_summary
# 引入 LegalRAG
from modules.legal_rag import LegalRAG

app = Flask(__name__)
# 增加请求大小限制为 50MB，支持大数据量导出
app.config['MAX_CONTENT_LENGTH'] = 50 * 1024 * 1024

# 初始化 RAG 系统 (全局单例)
print("正在初始化 LegalRAG 系统，请稍候...")
rag_system = LegalRAG()
print("LegalRAG 初始化完成")


@app.route('/')
def index():
    """主页"""
    available_types = get_available_types()
    # 依然获取类型用于展示统计信息，但在搜索中不再强制选择
    return render_template('index.html', case_types=available_types)


@app.route('/api/types')
def get_types():
    """获取可用的案件类型"""
    available_types = get_available_types()
    config = load_config()
    if not available_types:
        available_types = config.get('case_types', [])
    return jsonify({'types': available_types})


@app.route('/api/type-summary/<case_type>')
def get_type_summary(case_type):
    """获取指定类型的案例统计"""
    summary = get_cases_summary(case_type)
    return jsonify(summary)


@app.route('/api/search-stream', methods=['POST'])
def search_stream():
    """
    执行类案检索（SSE 流式返回进度）
    使用 LegalRAG 流程
    """
    data = request.get_json()
    query = data.get('query', '').strip()
    
    if not query:
        return jsonify({'error': '请输入查询内容'}), 400
    
    config = load_config()
    
    def generate():
        # 发送初始信息
        yield f"data: {json.dumps({'type': 'init', 'message': '正在分析您的查询...'})}\n\n"
        
        # 用于线程间通信的队列
        # 存放 (type, content) 元组
        # type: 'progress', 'result', 'error'
        msg_queue = queue.Queue()
        
        def progress_callback(message):
            msg_queue.put(('progress', message))
        
        # 在后台线程中执行 RAG 搜索
        def run_rag_search():
            try:
                print(f"开始 RAG 检索: {query}")
                # 传入回调函数获取真实进度
                results = rag_system.search(query, progress_callback=progress_callback)
                msg_queue.put(('result', results))
            except Exception as e:
                print(f"RAG search failed: {e}")
                msg_queue.put(('error', str(e)))
        
        search_thread = threading.Thread(target=run_rag_search)
        search_thread.start()
        
        # 循环读取队列消息直到结束
        while True:
            try:
                # 设置超时，防止死锁，同时也可以检查线程状态
                msg_type, content = msg_queue.get(timeout=0.5)
                
                if msg_type == 'progress':
                    yield f"data: {json.dumps({'type': 'progress', 'message': content})}\n\n"
                
                elif msg_type == 'result':
                    formatted_results = content
                    final_results = []
                    for idx, r in enumerate(formatted_results, 1):
                        final_results.append({
                            'rank': idx,
                            'filename': r.get('filename'),
                            'similarity_score': r.get('score'),
                            'summary': r.get('reason')[:100] + '...',
                            'reason': r.get('reason'),
                            'content': r.get('content'),
                            'header_info': r.get('header_info')
                        })
                    
                    final_data = {
                        'type': 'complete',
                        'results': final_results,
                        'message': f"检索完成，共找到 {len(final_results)} 个相关案例"
                    }
                    yield f"data: {json.dumps(final_data)}\n\n"
                    break # 完成后退出循环
                
                elif msg_type == 'error':
                    yield f"data: {json.dumps({'type': 'error', 'error': content})}\n\n"
                    break
                    
            except queue.Empty:
                if not search_thread.is_alive():
                    # 线程结束且队列为空，可能是异常退出或者逻辑漏洞
                    # 但如果有result或error应该已经break了
                    # 这里做双重保险，如果线程死了还没拿到结果，就退出
                    break
                # 继续等待
                continue

    return Response(generate(), mimetype='text/event-stream')


@app.route('/api/search', methods=['POST'])
def search():
    """
    执行类案检索（非流式版本）
    """
    data = request.get_json()
    query = data.get('query', '').strip()
    
    if not query:
        return jsonify({'error': '请输入查询内容'}), 400
    
    try:
        results = rag_system.search(query)
        
        final_results = []
        for idx, r in enumerate(results, 1):
            final_results.append({
                'rank': idx,
                'filename': r.get('filename'),
                'similarity_score': r.get('score'),
                'summary': r.get('reason')[:100] + '...',
                'reason': r.get('reason'),
                'content': r.get('content'),
                'header_info': r.get('header_info')
            })
            
        return jsonify({
            'results': final_results,
            'total': len(final_results)
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/export-excel', methods=['POST'])
def export_excel():
    """
    导出检索结果为 Excel 文件
    """
    # 支持 JSON 和表单两种提交方式
    if request.is_json:
        data = request.get_json()
    else:
        # 从表单数据解析 JSON
        form_data = request.form.get('data', '{}')
        data = json.loads(form_data)
    
    results = data.get('results', [])
    # 默认文件名前缀
    query_preview = "类案检索"
    
    if not results:
        return jsonify({'error': '没有可导出的结果'}), 400
    
    # 创建工作簿
    wb = Workbook()
    ws = wb.active
    ws.title = "检索结果"
    
    # 定义样式
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill(start_color="4A90E2", end_color="4A90E2", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    cell_alignment = Alignment(vertical="top", wrap_text=True)
    thin_border = Border(
        left=Side(style='thin', color='CCCCCC'),
        right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),
        bottom=Side(style='thin', color='CCCCCC')
    )
    
    # 设置表头
    headers = ['排名', '文件名', '相似度', '相似理由', '全文内容']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # 设置列宽
    ws.column_dimensions['A'].width = 8   # 排名
    ws.column_dimensions['B'].width = 30  # 文件名
    ws.column_dimensions['C'].width = 10  # 相似度
    ws.column_dimensions['D'].width = 60  # 相似理由
    ws.column_dimensions['E'].width = 80  # 全文内容
    
    # 填充数据
    for row, result in enumerate(results, 2):
        ws.cell(row=row, column=1, value=result.get('rank', row-1)).alignment = Alignment(horizontal="center", vertical="top")
        ws.cell(row=row, column=2, value=result.get('filename', '')).alignment = cell_alignment
        ws.cell(row=row, column=3, value=result.get('similarity_score', '')).alignment = Alignment(horizontal="center", vertical="top")
        ws.cell(row=row, column=4, value=result.get('reason', '')).alignment = cell_alignment
        ws.cell(row=row, column=5, value=result.get('content', '')).alignment = cell_alignment
        
        # 添加边框
        for col in range(1, 6):
            ws.cell(row=row, column=col).border = thin_border
    
    # 冻结首行
    ws.freeze_panes = 'A2'
    
    # 保存到内存
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    # 生成文件名
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f'{query_preview}_{timestamp}.xlsx'
    
    # 使用 urllib 编码中文文件名
    from urllib.parse import quote
    encoded_filename = quote(filename)
    
    response = Response(
        output.getvalue(),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response.headers['Content-Disposition'] = f"attachment; filename*=UTF-8''{encoded_filename}"
    response.headers['X-Filename'] = encoded_filename
    response.headers['Access-Control-Expose-Headers'] = 'X-Filename, Content-Disposition'
    return response


if __name__ == '__main__':
    print("=" * 50)
    print("行政处罚决定书类案检索系统 (Powered by LegalRAG)")
    print("=" * 50)
    
    print("\n启动服务...")
    print("访问地址: http://localhost:5001")
    print("=" * 50)
    
    app.run(debug=True, host='0.0.0.0', port=5001, threaded=True)
